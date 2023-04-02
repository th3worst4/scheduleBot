from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime

wb = load_workbook('..\Cronograma.xlsx')
ws = wb.active

def today():
    x = int(datetime.datetime.now().strftime('%w')) + 1
    if x == 1:
        return "sunday"
    else:
        info = list()
        for row in range(2,15):
            content = ws[get_column_letter(x) + f'{row}'].value
            info.append(str(ws['A'+f'{row}'].value) + ": " + str(content))
        return info

def anyday(command):
    info = list()
    
    weekdays = ['/sunday', '/monday', '/tuesday', '/wednesday', '/thursday', '/friday', '/saturday']
    x = weekdays.index(command) + 1
    
    for row in range(2,15):
        content = ws[get_column_letter(x) + f'{row}'].value
        info.append(str(ws['A'+f'{row}'].value) + ": " + str(content))
    return info

def tomorrow():
    info = list()
    
    x = int(datetime.datetime.now().strftime('%w')) + 1
    
    if x == 7:
        x = 0
        return "sunday"
    else:
        x+=1
    
    for row in range(2,15):
        content = ws[get_column_letter(x) + f'{row}'].value
        info.append(str(ws['A'+f'{row}'].value) + ": " + str(content))
    return info